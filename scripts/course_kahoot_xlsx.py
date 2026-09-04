# -*- coding: utf-8 -*-
"""把講義的章末小考轉成 Kahoot 可匯入的 xlsx（每次上課一份，兩章共 20 題）。

題庫用的是既有的 `public/content/works/{slug}/quizzes/`，不另出一套題——
題目改了重跑就好。成品依 docs/repo-hygiene.md 不進 git，輸出到 Drive：
  G:\\我的雲端硬碟\\資料\\知識圖工作室\\教學\\{課程資料夾}\\Kahoot\\

選項順序直接沿用題庫，**這裡不再另外打散**——打散已由
`scripts/course_quiz_shuffle.py` 一次做在題庫原檔上（原本正解位置是固定
循環，三門課十六章序列完全相同）。兩邊共用同一個順序，紙本與 Kahoot 才
對得起來；若在這裡再洗一次，兩者就會不一致。

用法：
  python scripts/course_kahoot_xlsx.py                 # 三門課全出
  python scripts/course_kahoot_xlsx.py --course=ch 1 2 # 指定課程與次數
"""
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parent.parent
WORKS = ROOT / 'public' / 'content' / 'works'
DRIVE = Path(r'G:\我的雲端硬碟\資料\知識圖工作室\教學')

COURSES = {
    'wr': ('115-1_世界宗教文化導論', 'world-religions-intro', 'wr2', '世界宗教文化導論'),
    'sl': ('宗教系國文講義', 'sinographic-literature', 'sl1', '宗教系國文講義'),
    'ch': ('115-1_基督宗教概論', 'christianity-intro', 'ch1', '基督宗教概論'),
}

# Kahoot 的字數上限各處說法不一（120/75 與 95/60 兩種），取嚴的那組，
# 兩邊都能過。超過的題目不截斷、只列出來人工改——截斷會把題意切壞。
Q_MAX, A_MAX = 95, 60
TIME_LIMIT = 30          # 允許值：5, 10, 20, 30, 60, 90, 120, 240


def strip(html):
    return re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', '', html)).strip()


def parse_quiz(path):
    """讀一章的小考，回傳 [(題幹, [四個選項], 正解索引)]。"""
    s = path.read_text(encoding='utf-8')
    blocks = re.findall(r'<li>(.*?)<ul class="q-options">(.*?)</ul>', s, re.S)
    m = re.search(r'<div class="quiz-answers">.*?<ol>(.*?)</ol>', s, re.S)
    keys = re.findall(r'<strong>\(([A-D])\)</strong>', m.group(1)) if m else []
    if len(keys) != len(blocks):
        raise ValueError(f'{path.name}：題數 {len(blocks)} 與答案數 {len(keys)} 不符')
    out = []
    for (stem, opts_html), key in zip(blocks, keys):
        opts = [re.sub(r'^\([A-D]\)\s*', '', strip(o))
                for o in re.findall(r'<li>(.*?)</li>', opts_html, re.S)]
        out.append((strip(stem), opts, 'ABCD'.index(key)))
    return out


def build(course, lesson):
    folder, slug, prefix, title = COURSES[course]
    chapters = (lesson * 2 - 1, lesson * 2)
    rows, warn = [], []
    for ch in chapters:
        f = WORKS / slug / 'quizzes' / f'{prefix}-ch{ch:02d}.html'
        if not f.exists():
            continue
        for i, (stem, opts, correct) in enumerate(parse_quiz(f), 1):
            if len(stem) > Q_MAX:
                warn.append(f'第{ch}章第{i}題　題幹 {len(stem)} 字')
            for o in opts:
                if len(o) > A_MAX:
                    warn.append(f'第{ch}章第{i}題　選項 {len(o)} 字')
            rows.append((stem, opts, correct))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Kahoot'
    # 前八列照官方範本留給說明與標題，資料自第 9 列起；欄序 B 題幹、C–F 選項、
    # G 秒數、H 正解。若 Kahoot 的匯入器挑格式，直接把 B9:H 這塊貼進它自己
    # 下載的範本即可——欄序一樣。
    ws['B1'] = f'{title}　第 {lesson} 次（第 {chapters[0]}、{chapters[1]} 章）Kahoot 題庫'
    ws['B1'].font = Font(bold=True, size=14)
    ws['B2'] = '由講義章末小考自動轉出（scripts/course_kahoot_xlsx.py）；順序與紙本小考一致。'
    ws['B3'] = f'題幹上限 {Q_MAX} 字、選項上限 {A_MAX} 字、每題 {TIME_LIMIT} 秒。'
    heads = ['', '題目（Question）', '選項 1', '選項 2', '選項 3', '選項 4',
             '秒數（Time limit）', '正解（Correct answer(s)）']
    for col, h in enumerate(heads, 1):
        c = ws.cell(row=8, column=col, value=h)
        c.font = Font(bold=True)
    for n, (stem, opts, correct) in enumerate(rows, 1):
        r = 8 + n
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=stem).alignment = Alignment(wrap_text=True)
        for j, o in enumerate(opts):
            ws.cell(row=r, column=3 + j, value=o).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=7, value=TIME_LIMIT)
        ws.cell(row=r, column=8, value=correct + 1)
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 58
    for col in 'CDEF':
        ws.column_dimensions[col].width = 34
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10

    outdir = DRIVE / folder / 'Kahoot'
    outdir.mkdir(parents=True, exist_ok=True)
    out = outdir / f'第{lesson}次_第{chapters[0]}-{chapters[1]}章.xlsx'
    wb.save(out)
    return out, len(rows), warn


if __name__ == '__main__':
    sys.stdout.reconfigure(encoding='utf-8')
    args = sys.argv[1:]
    picked = next((a.split('=')[1] for a in args if a.startswith('--course=')), None)
    lessons = [int(a) for a in args if not a.startswith('--')] or list(range(1, 9))
    for course in ([picked] if picked else list(COURSES)):
        for n in lessons:
            out, cnt, warn = build(course, n)
            print(f'✔ {out}　（{cnt} 題）')
            for w in warn:
                print(f'   ⚠ 超出字數上限，請人工改短：{w}')
