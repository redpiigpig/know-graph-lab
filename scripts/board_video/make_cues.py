# -*- coding: utf-8 -*-
"""把解說腳本切成 cue，產出 cues.json（給黑板渲染用）與細流表.xlsx（給人看的）。

用法：
    python make_cues.py                # 用估算語速
    python make_cues.py --cps 4.2      # 調語速（字/秒）

配圖建議來自 素材/預告片/建議配圖.json（節點 → 候選截圖清單），沒有就留空。
"""
import argparse
import json
import math
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import spec_chiikawa as spec  # noqa: E402

PROJ = Path(r"G:\我的雲端硬碟\創作\影片創作\人魚島解說")
SCRIPT = PROJ / "腳本.txt"
SUGGEST_JSON = PROJ / "素材" / "預告片" / "建議配圖.json"

SENT_END = "。！？"
SOFT_BREAK = "，、；："
MAX_CUE = 44          # 一條 cue 的字數上限
MERGE_UNDER = 26      # 短句合併門檻


def split_sentences(text: str) -> list[str]:
    """先按句末標點切，再把過長的句子在逗號處拆開，最後把過短的併回去。"""
    out, buf = [], ""
    for ch in text:
        buf += ch
        if ch in SENT_END:
            out.append(buf.strip())
            buf = ""
    if buf.strip():
        out.append(buf.strip())

    pieces = []
    for s in out:
        while len(s) > MAX_CUE:
            cut = -1
            for i in range(MAX_CUE, max(MAX_CUE - 18, 8), -1):
                if s[i - 1] in SOFT_BREAK:
                    cut = i
                    break
            if cut < 0:
                cut = MAX_CUE
            pieces.append(s[:cut].strip())
            s = s[cut:].strip()
        if s:
            pieces.append(s)

    merged = []
    for p in pieces:
        if merged and len(merged[-1]) < MERGE_UNDER and len(merged[-1]) + len(p) <= MAX_CUE:
            merged[-1] += p
        else:
            merged.append(p)
    return merged


def strip_marker(text: str) -> tuple[str, str | None]:
    """把段首的【第 N 幕：…】拿掉，回傳 (台詞, 幕名)。"""
    m = re.match(r"^【([^】]+)】\s*", text)
    if m:
        return text[m.end():].strip(), m.group(1)
    return text, None


def build():
    ap = argparse.ArgumentParser()
    ap.add_argument("--cps", type=float, default=4.5, help="估算語速（字/秒）")
    ap.add_argument("--pause", type=float, default=0.45, help="每條 cue 之後的停頓（秒）")
    args = ap.parse_args()

    lines = SCRIPT.read_text(encoding="utf-8").split("\n")
    nodes = {n["id"]: n for n in spec.NODES}
    suggest = json.loads(SUGGEST_JSON.read_text(encoding="utf-8")) if SUGGEST_JSON.exists() else {}

    # 先蒐集每個節點被哪些 cue 用到，才能把 bullet 平均分配給這些 cue
    raw = []  # (para_idx, node_id, chapter, text)
    for i, line in enumerate(lines):
        node_id = spec.PARA_NODE.get(i)
        if not node_id:
            continue
        text, inline_chapter = strip_marker(line.strip())
        chapter = spec.CHAPTERS.get(i) or inline_chapter
        for s in split_sentences(text):
            raw.append([i, node_id, chapter, s])
            chapter = None  # 幕名只掛在該幕第一條

    per_node_total, per_node_seen = {}, {}
    for _, nid, _, _ in raw:
        per_node_total[nid] = per_node_total.get(nid, 0) + 1

    cues, t = [], 0.0
    prev_node = None
    for idx, (para, nid, chapter, text) in enumerate(raw):
        node = nodes[nid]
        seen = per_node_seen.get(nid, 0)
        per_node_seen[nid] = seen + 1
        total = per_node_total[nid]
        nb = len(node["bullets"])
        # 用 ceil：節點一登場就至少寫出第一點，否則會出現「有框沒字」的空板
        reveal = min(nb, max(1, math.ceil((seen + 1) / total * nb))) if total else nb
        dur = round(max(1.8, len(text) / args.cps + args.pause), 2)
        cues.append(dict(
            i=idx, t=round(t, 2), dur=dur, text=text, node=nid,
            reveal=reveal, chapter=chapter, para=para,
            first_visit=(seen == 0),
            travel=(prev_node is not None and prev_node != nid),
            group=node["group"],
        ))
        t += dur
        prev_node = nid

    total_s = round(t, 1)
    out = dict(
        board=spec.BOARD,
        nodes=spec.NODES,
        edges=[dict(a=a, b=b, label=l) for a, b, l in spec.EDGES],
        cues=cues,
        total=total_s,
        cps=args.cps,
    )
    (PROJ / "cues.json").write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")

    write_xlsx(cues, nodes, suggest, total_s)
    print(f"cue 數 {len(cues)}／全片估算 {int(total_s//60)} 分 {int(total_s%60)} 秒")
    return out


def mmss(s: float) -> str:
    return f"{int(s // 60):02d}:{int(s % 60):02d}"


def write_xlsx(cues, nodes, suggest, total_s):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "細流表"
    ws.append(["時間", "台詞", "影片/配圖", "音效", "背景音樂"])
    head = ws[1]
    for c in head:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="1F3B4D")
        c.font = Font(bold=True, color="FFFFFF")

    band = PatternFill("solid", fgColor="E8EEF4")
    bgm_by_group = {
        "序幕": "開場輕快（可愛系）",
        "劇情": "敘事底噪，戰鬥段落轉緊張",
        "人魚島": "懸疑、低音鋪底",
        "倫理": "思辨、極簡鋼琴",
        "自然": "低沉、弦樂",
        "神話": "古樂／史詩感",
        "假設": "思辨、極簡鋼琴",
        "結語": "溫暖收束",
    }

    for cue in cues:
        node = nodes[cue["node"]]
        if cue["chapter"]:
            ws.append([f"── {cue['chapter']} ──", "", "", "", ""])
            for c in ws[ws.max_row]:
                c.fill = band
                c.font = Font(bold=True)

        pic = [f"黑板：{node['title']}（顯示到第 {cue['reveal']} 點）"]
        asset = node.get("asset")
        if asset and asset.startswith("【自備】"):
            pic.append(asset)
        elif asset:
            pd = spec.PD_IMAGES.get(asset)
            if pd:
                pic.append(f"公有領域圖：{pd['title']}")
        if cue["first_visit"]:
            hint = suggest.get(cue["node"])
            if hint:
                pic.append("建議預告片截圖：" + "／".join(hint.get("shots", [])))
                if hint.get("note"):
                    pic.append(hint["note"])

        sfx = ""
        if cue["chapter"]:
            sfx = "轉場：粉筆擦板／書寫聲"
        elif cue["travel"]:
            sfx = "鏡頭移動：短暫 whoosh"

        row = [f"{mmss(cue['t'])}–{mmss(cue['t'] + cue['dur'])}", cue["text"],
               "\n".join(pic), sfx, bgm_by_group.get(cue["group"], "")]
        ws.append(row)
        for c in ws[ws.max_row]:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    ws.append([mmss(total_s), "（全片估算長度）", "", "", ""])
    for col, w in zip("ABCDE", (14, 52, 46, 24, 22)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(PROJ / "人魚島解說_細流表.xlsx")


if __name__ == "__main__":
    build()
