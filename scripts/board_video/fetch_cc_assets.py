# -*- coding: utf-8 -*-
"""抓可用授權的素材：CC 背景音樂／音效 ＋ 公有領域圖，全部不需 API 金鑰。

來源分工（都會把授權與出處記下來）：
  音樂  ccMixter（CC BY／BY-NC，人聲少、可挑 instrumental）＋ archive.org netlabels
  音效  Openverse（背後多半是 Freesound 的 CC0／CC BY），匿名額度低，所以慢速重試
  圖片  Wikimedia Commons（搜尋或分類），只收公有領域／CC

輸出：素材索引.xlsx（分類／檔名／授權／建議用在哪一段）＋ 授權標示.txt（貼 YouTube 說明欄）。

用法：
    python fetch_cc_assets.py                 # 全抓
    python fetch_cc_assets.py --only 音樂     # 只抓某一類（音樂／音效／圖片）
"""
import argparse
import json
import re
import time
import urllib.parse
import urllib.request
from pathlib import Path

PROJ = Path(r"G:\我的雲端硬碟\創作\影片創作\人魚島解說")
MUSIC, SFX, PD = PROJ / "素材" / "音樂", PROJ / "素材" / "音效", PROJ / "素材" / "公有領域"
UA = {"User-Agent": "know-graph-lab-video/1.0 (personal explainer video project)"}

# 分類 → (ccMixter 標籤, archive.org 關鍵詞, 要幾首, 建議用在哪一段)
MUSIC_PLAN = {
    "01_開場輕快": ("ukulele,happy", "playful light acoustic", 4, "序幕：多馬茶房開場、片名卡"),
    "02_敘事鋪底": ("ambient,instrumental", "calm background instrumental", 5, "第一～三幕：劇情複述"),
    "03_緊張戰鬥": ("drums,percussion", "tension percussion action", 4, "第四幕：超辣咖哩決戰"),
    "04_黑暗懸疑": ("dark,ambient", "dark ambient suspense", 5, "第五幕黑暗反轉、第九幕大自然反撲"),
    "05_思辨鋼琴": ("piano,minimal", "solo piano minimal", 5, "第六～八幕倫理分析、共業假設"),
    "06_神話史詩": ("orchestral,cinematic", "orchestral cinematic epic", 5, "第十～十二幕：米諾陶洛斯、八岐大蛇、生民"),
    "07_溫暖收束": ("acoustic,guitar", "warm gentle acoustic", 4, "結語：擁抱不適感"),
}

SFX_PLAN = {
    "粉筆書寫": ("chalk writing", 3, "黑板寫字、重點浮現"),
    "擦黑板": ("blackboard eraser", 2, "換幕轉場"),
    "whoosh轉場": ("whoosh transition", 4, "鏡頭在黑板上飛移"),
    "叮提示": ("ding bell short", 3, "關鍵詞、標記重點"),
    "重擊": ("impact boom cinematic", 3, "揭露真相、轉折"),
    "心跳": ("heartbeat", 2, "懸疑鋪陳"),
    "海浪": ("ocean waves", 2, "人魚島場景"),
    "人群歡呼": ("crowd cheering", 2, "營火晚會、慶功"),
    "翻頁": ("page turn paper", 2, "切換神話案例"),
}

# 圖片：前綴 → (Commons 搜尋詞清單, 分類清單, 幾張, 用在哪)
PD_PLAN = {
    "賽蓮": (["Odysseus sirens vase", "siren mythology painting"],
             ["Sirens in art"], 4, "N21 賽蓮｜強者"),
    "米諾陶洛斯": (["Theseus Minotaur ancient vase"], ["Minotaur"], 4, "N52 米諾陶洛斯"),
    "八岐大蛇": (["Susanoo Yamata no Orochi ukiyo-e"], ["Yamata no Orochi"], 4, "N53 八岐大蛇"),
    "普羅米修斯": (["Prometheus bound painting"], ["Prometheus"], 4, "N54 竊取神力的人"),
    "詩經生民": (["Shijing Book of Odes page", "詩經"], ["Shijing"], 4, "N55 詩經・大雅・生民"),
    "八百比丘尼": (["Yao Bikuni", "八百比丘尼", "Ningyo Japanese mermaid"],
                   ["Ningyo"], 4, "N51 八百比丘尼"),
    "克里特迷宮": (["Knossos labyrinth Crete"], ["Labyrinths"], 3, "N52 迷宮"),
    "人魚圖像": (["mermaid medieval manuscript"], ["Mermaids in art"], 3, "人魚／永生設定"),
}


def get(url: str, tries=3, wait=8.0):
    for k in range(tries):
        try:
            req = urllib.request.Request(url, headers=UA)
            with urllib.request.urlopen(req, timeout=45) as r:
                return json.loads(r.read().decode("utf-8", "replace"))
        except Exception as e:
            if k == tries - 1:
                print(f"    查詢失敗：{e}")
                return None
            time.sleep(wait * (k + 1))
    return None


def download(url: str, dest: Path, min_bytes=3000) -> bool:
    if dest.exists() and dest.stat().st_size > min_bytes:
        return True
    try:
        head = dict(UA)
        if "ccmixter" in url:      # ccMixter 直接抓檔會 403，要帶 Referer
            head["Referer"] = "https://ccmixter.org/"
        req = urllib.request.Request(url, headers=head)
        with urllib.request.urlopen(req, timeout=180) as r:
            data = r.read()
        if len(data) < min_bytes:
            return False
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(data)
        return True
    except Exception as e:
        print(f"    下載失敗 {url[:70]}… {e}")
        return False


def safe(name: str, limit=46) -> str:
    return (re.sub(r"[\\/:*?\"<>|\r\n\t]", "_", name).strip()[:limit] or "untitled").strip("_ ")


# ── 音樂：ccMixter ──────────────────────────────────────────────
def ccmixter(tags: str, want: int):
    url = ("https://ccmixter.org/api/query?"
           + urllib.parse.urlencode({"f": "json", "tags": tags, "limit": want * 3,
                                     "sinced": "5 years ago", "sort": "rank"}))
    data = get(url)
    out = []
    for r in data or []:
        files = r.get("files") or []
        mp3 = next((f for f in files
                    if (f.get("download_url", "").lower().endswith((".mp3", ".ogg")))), None)
        if not mp3:
            continue
        lic = r.get("license_name") or "見來源頁"
        out.append(dict(title=r.get("upload_name") or "untitled", creator=r.get("user_name") or "unknown",
                        license=lic, version="", source=r.get("file_page_url") or r.get("upload_page_url", ""),
                        url=mp3["download_url"], provider="ccMixter",
                        duration=mp3.get("file_length") or ""))
        if len(out) >= want:
            break
    return out


# ── 音樂：archive.org（netlabels，授權欄位可機讀）─────────────────
def archive_audio(keywords: str, want: int):
    q = (f'({keywords}) AND mediatype:(audio) AND collection:(netlabels) '
         f'AND licenseurl:(*creativecommons.org*)')
    url = ("https://archive.org/advancedsearch.php?"
           + urllib.parse.urlencode({"q": q, "rows": want * 2, "page": 1, "output": "json",
                                     "fl[]": "identifier"})
           + "&fl%5B%5D=title&fl%5B%5D=creator&fl%5B%5D=licenseurl")
    data = get(url)
    docs = ((data or {}).get("response") or {}).get("docs") or []
    out = []
    for d in docs:
        ident = d.get("identifier")
        meta = get(f"https://archive.org/metadata/{ident}")
        if not meta:
            continue
        files = [f for f in meta.get("files", [])
                 if f.get("name", "").lower().endswith(".mp3") and int(f.get("size") or 0) > 500_000]
        if not files:
            continue
        f = files[0]
        lic = d.get("licenseurl", "")
        out.append(dict(title=f"{d.get('title', ident)} - {f['name']}"[:70],
                        creator=(d.get("creator") or "unknown") if isinstance(d.get("creator"), str)
                        else "、".join(d.get("creator") or ["unknown"]),
                        license=re.sub(r".*creativecommons\.org/licenses/", "CC ", lic).replace("/", " ").upper()
                        or "CC（見來源）", version="",
                        source=f"https://archive.org/details/{ident}",
                        url=f"https://archive.org/download/{ident}/{urllib.parse.quote(f['name'])}",
                        provider="archive.org",
                        duration=round(float(f.get("length") or 0) if str(f.get("length", "")).replace(".", "").isdigit() else 0, 1)))
        if len(out) >= want:
            break
        time.sleep(0.4)
    return out


# ── 音效：Openverse（匿名額度低，慢慢來）─────────────────────────
def openverse_audio(query: str, want: int, max_dur=15.0):
    url = ("https://api.openverse.org/v1/audio/?"
           + urllib.parse.urlencode({"q": query, "license": "cc0,by", "page_size": 30}))
    data = get(url, tries=4, wait=20)
    out = []
    for r in (data or {}).get("results", []):
        dur = (r.get("duration") or 0) / 1000.0
        if dur > max_dur or not r.get("url"):
            continue
        out.append(dict(title=r.get("title") or "untitled", creator=r.get("creator") or "unknown",
                        license=(r.get("license") or "").upper(), version=r.get("license_version") or "",
                        source=r.get("foreign_landing_url") or r["url"], url=r["url"],
                        provider=r.get("provider") or "openverse", duration=round(dur, 1)))
        if len(out) >= want:
            break
    return out


# ── 圖片：Wikimedia Commons ────────────────────────────────────
def commons(params: dict, want: int):
    base = {"action": "query", "format": "json", "prop": "imageinfo",
            "iiprop": "url|extmetadata", "iiurlwidth": "1600"}
    data = get("https://commons.wikimedia.org/w/api.php?" + urllib.parse.urlencode({**base, **params}))
    out = []
    for page in ((data or {}).get("query", {}).get("pages", {}) or {}).values():
        info = (page.get("imageinfo") or [{}])[0]
        if not info.get("thumburl"):
            continue
        meta = info.get("extmetadata", {}) or {}
        lic = (meta.get("LicenseShortName", {}).get("value") or "見來源頁").strip()
        if "fair use" in lic.lower() or "non-free" in lic.lower():
            continue
        artist = re.sub(r"<[^>]+>", "", meta.get("Artist", {}).get("value") or "unknown").strip()
        out.append(dict(title=page.get("title", "").replace("File:", ""), creator=artist or "unknown",
                        license=lic, version="", source=info.get("descriptionurl", ""),
                        url=info["thumburl"], provider="Wikimedia Commons", duration=""))
        if len(out) >= want:
            break
    return out


def commons_images(searches, cats, want):
    out, seen = [], set()
    for s in searches:
        for r in commons({"generator": "search", "gsrsearch": f"{s} filetype:bitmap",
                          "gsrnamespace": "6", "gsrlimit": str(want * 3)}, want):
            if r["title"] not in seen:
                seen.add(r["title"]); out.append(r)
        if len(out) >= want:
            return out[:want]
    for c in cats:
        for r in commons({"generator": "categorymembers", "gcmtitle": f"Category:{c}",
                          "gcmtype": "file", "gcmlimit": str(want * 3)}, want):
            if r["title"] not in seen:
                seen.add(r["title"]); out.append(r)
        if len(out) >= want:
            break
    return out[:want]


def ext_of(url: str, default=".mp3") -> str:
    return Path(urllib.parse.urlparse(url).path).suffix or default


def run(only=None):
    rows = []
    if only in (None, "音樂"):
        print("── 背景音樂 ──")
        for cat, (tags, kw, want, usage) in MUSIC_PLAN.items():
            got = ccmixter(tags, want)
            if len(got) < want:
                got += archive_audio(kw, want - len(got))
            print(f"  {cat}: {len(got)} 首")
            for i, r in enumerate(got, 1):
                dest = MUSIC / cat / f"{i:02d}_{safe(r['title'])}{ext_of(r['url'])}"
                if download(r["url"], dest, min_bytes=200_000):
                    rows.append(dict(kind="音樂", cat=cat, path=dest, usage=usage, **r))
            time.sleep(1)

    if only in (None, "音效"):
        print("── 音效（Openverse 匿名額度低，慢慢抓）──")
        for cat, (q, want, usage) in SFX_PLAN.items():
            got = openverse_audio(q, want)
            print(f"  {cat}: {len(got)} 個")
            for i, r in enumerate(got, 1):
                dest = SFX / cat / f"{i:02d}_{safe(r['title'])}{ext_of(r['url'])}"
                if download(r["url"], dest, min_bytes=8000):
                    rows.append(dict(kind="音效", cat=cat, path=dest, usage=usage, **r))
            time.sleep(12)

    if only in (None, "圖片"):
        print("── 公有領域圖 ──")
        for prefix, (searches, cats, want, usage) in PD_PLAN.items():
            got = commons_images(searches, cats, want)
            print(f"  {prefix}: {len(got)} 張")
            for i, r in enumerate(got, 1):
                dest = PD / f"{prefix}_{i:02d}{ext_of(r['url'], '.jpg')}"
                if download(r["url"], dest, min_bytes=20000):
                    rows.append(dict(kind="圖片", cat=prefix, path=dest, usage=usage, **r))
            time.sleep(0.6)

    merge_and_write(rows)
    print(f"\n本輪 {len(rows)} 件")


def merge_and_write(rows):
    """跟上一輪的紀錄合併，避免分批執行時把先前抓到的蓋掉。"""
    ledger = PROJ / "素材" / "_ledger.json"
    old = json.loads(ledger.read_text(encoding="utf-8")) if ledger.exists() else []
    seen = {r["file"] for r in old}
    for r in rows:
        if r["path"].name in seen:
            continue
        old.append(dict(kind=r["kind"], cat=r["cat"], file=r["path"].name,
                        rel=str(r["path"].relative_to(PROJ)), title=r["title"], creator=r["creator"],
                        license=r["license"], version=r["version"], duration=r["duration"],
                        usage=r["usage"], source=r["source"], provider=r["provider"]))
        seen.add(r["path"].name)
    old = [r for r in old if (PROJ / r["rel"]).exists()]
    ledger.write_text(json.dumps(old, ensure_ascii=False, indent=1), encoding="utf-8")
    write_index(old)
    write_credits(old)


def lic_text(r) -> str:
    if r["license"] in ("BY", "CC0"):
        return f"CC {r['license']} {r['version']}".strip()
    return r["license"]


def write_index(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    wb = Workbook()
    ws = wb.active
    ws.title = "素材索引"
    ws.append(["類型", "分類", "檔名", "標題", "作者", "授權", "長度(秒)", "建議用在", "來源網址"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in sorted(rows, key=lambda x: (x["kind"], x["cat"], x["file"])):
        ws.append([r["kind"], r["cat"], r["file"], r["title"], r["creator"], lic_text(r),
                   r["duration"], r["usage"], r["source"]])
        for c in ws[ws.max_row]:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCDEFGHI", (8, 16, 34, 30, 18, 18, 10, 32, 46)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(PROJ / "素材索引.xlsx")


def write_credits(rows):
    lines = ["【素材授權標示】貼進 YouTube 說明欄。CC BY 必須保留作者與授權連結；",
             "CC0／公有領域不強制，但一併列出比較乾淨。", ""]
    for kind in ("音樂", "音效", "圖片"):
        sub = [r for r in rows if r["kind"] == kind]
        if not sub:
            continue
        lines.append(f"■ {kind}")
        for r in sorted(sub, key=lambda x: x["cat"]):
            lines.append(f"・{r['title']} — {r['creator']}（{lic_text(r)}）{r['source']}")
        lines.append("")
    lines += ["■ 影片畫面",
              "・《劇場版 吉伊卡哇 人魚島的秘密》預告片畫面（東宝／Medialink 官方預告）：評論用途引用。",
              "・《魔法公主》預告片畫面（Studio Ghibli／東宝／GKIDS 官方預告）：評論用途引用。"]
    (PROJ / "授權標示.txt").write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", choices=["音樂", "音效", "圖片"])
    run(ap.parse_args().only)
